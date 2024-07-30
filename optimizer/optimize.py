from optimizer import constant
from optimizer import utils
from ortools.sat.python import cp_model
#from pyscipopt import *
from optimizer.master import masterClass
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from flaretool.holiday import JapaneseHolidays
from collections import defaultdict

import pandas as pd
import os
import datetime
import pprint
import logging
import math

debug = 1

class solverClass:
    
    #初期化
    def __init__(self, params : dict, master : masterClass, logger : logging.Logger) -> None:        
        self.params = params
        self.master = master
        self.relax_level = 0
        self.vars = {}
        self.coeffs = {}
        self.prob = None
        self.solver = None
        self.bestSol = None
        self.solver_engin_is = "sat"
        self.month_list = []
        self.offset_day = 0
        self.offset_month = 0
        self.nday_this_month = 0
        self.fiscal_year_start = 0
        self.logger = logger
        self.is_legal_holiday = None
        self.is_relax_log = False
        self.relax_sol = defaultdict(dict)
        self.constraints = []

        self.accept_master()

        pass
    
    #解取得
    def getSol(self, var) -> float:
        if(var is None):
            return 0.0
        if(self.solver_engin_is == "scip"):
            return self.bestSol[var]
        else:
            return self.solver.Value(var)

    #初期化
    def init(self) -> None:
        #変数と目的関数の係数をクリア
        self.vars.clear()
        self.coeffs.clear()
        self.constraints.clear()
        #モデルインスタンス生成
        if(self.solver_engin_is == "scip"):
            #self.prob = Model()
            pass
        else:
            self.prob = cp_model.CpModel()
        return 

    #結果チェック
    def output_check(self) -> None:
        
        master = self.master
        nday = self.params[constant.param_nday]['value']
        skill = master.skill

        for worker_code, routes in skill.items():
            for day in range(0, nday):
                var = self.get_rankvar(worker_code, day)
                sol = self.getSol(var)
                if(sol > 0.1):
                    print(f"{worker_code},{day},{sol}")

        return 

    #結果出力
    def output(self, filepath : str) -> None:
        
        #self.output_check()

        dir_name = os.path.dirname(filepath)
        if(not os.path.exists(dir_name)):
            os.mkdir(dir_name)

        dfs = dict()

        #SHEET1. シフトシスト出力
        headers = ["従業員コード", "ルート", "日付", "曜日", "休日希望かどうか", "他業務", "スキル", "対象部署"]
        sht_name1 = "シフトリスト"

        master : masterClass = self.master
        start  : datetime    = self.params[constant.param_targetmonth]["value"]
        nday = self.params[constant.param_nday]["value"]
        dep  = self.params[constant.param_department]["value"]
        worker_codes = master.worker_code
        worker_names = master.worker_name
        skills = master.skill
        route = master.route
        holidays = master.holiday
        otherworks = master.otherwork
        ignore = master.ignores

        # 割り付け結果出力
        data = [ ]
        for worker_code, routes in skills.items():
            if not worker_code in holidays:
                continue
            for day in range(nday):
                date  = start + datetime.timedelta(day)
                week  = date.strftime('%A')
                holi  = holidays[worker_code][day]
                other = otherworks[worker_code][day]

                if(worker_code in ignore and other != ""):
                    data.append([worker_code, other, f"{date:%Y/%m/%d}", week, "シフト割付対象外作業員", other, "", dep])
                    continue

                for route_code, skill_level in routes.items():
                    if(skill_level == 0):
                        continue
                    var = self.get_deltavar(worker_code, route_code, day)
                    if(var is not None):
                        if(self.getSol(var) > 0.1):
                            data.append([worker_code, route_code, f"{date:%Y/%m/%d}", week, "" if holi == "" else holi, other, skill_level, dep])
        dfs[sht_name1] = pd.DataFrame(data, columns = headers)

        #SHEET2. シフトテーブル出力
        headers = ["従業員CD", "従業員"] + [ (start + datetime.timedelta(i)).strftime('%Y/%m/%d') + (start + datetime.timedelta(i)).strftime('%A')[0:3] for i in range(nday) ]
        sht_name2 = "シフトテーブル"
        
        data = [ ]
        for worker_code, routes in skills.items():
            worker_name = worker_names[worker_code]
            res = [worker_code, worker_name]
            if not worker_code in holidays:
                continue       

            res += ["" for day in range(nday)]     
            for day in range(nday):                
                for route_code, skill_level in routes.items():
                    var = self.get_deltavar(worker_code, route_code, day)
                    if(var is not None):
                        if(self.getSol(var) > 0.1):
                            res[day + 2] = route_code
            data.append(res)

        dfs[sht_name2] = pd.DataFrame(data, columns = headers)

        #SHEET3. 作業時間テーブル出力
        headers = ["従業員CD", "従業員"] + [ (start + datetime.timedelta(i)).strftime('%Y/%m/%d') for i in range(nday) ]
        sht_name3 = "作業時間テーブル"
        data = [ ]
        for worker_code, routes in skills.items():
            worker_name = worker_names[worker_code]
            res = [worker_code, worker_name]
            if not worker_code in holidays:
                continue       

            res += ["" for day in range(nday)]

            for day in range(nday):
                for route_code, skill_level in routes.items():
                    if(not route_code in route):
                        continue
                    times = route[route_code][day]
                    if(int(times[1][0]) < int(times[0][0])):
                        dt = 24 * 60 + int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                    else:
                        dt = int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                    var = self.get_deltavar(worker_code, route_code, day)
                    if(var is not None):
                        if(self.getSol(var) > 0.1):
                            res[day + 2] = dt
            data.append(res)
        dfs[sht_name3] = pd.DataFrame(data, columns = headers)

        #SHEET4. エラーリスト出力
        headers = ["日付", "従業員コード", "制約名称"]
        sht_name4 = "エラーリスト"
        data = [ ]
        
        for const_info in self.constraints:
            name = const_info.get("name", "")
            is_relax = const_info.get("is_relax", False)
            lhs = const_info.get("lhs", [])
            rhs = const_info.get("rhs", [])
            keys = const_info.get("keys", [])
            if(is_relax and self.checkConstraintViolation(lhs, rhs)):
                worker_code = keys.get("worker_code", "")
                day = keys.get("day", "")
                data.append([start + datetime.timedelta(day), worker_code, name])

        dfs[sht_name4] = pd.DataFrame(data, columns = headers) 

        #SHEET5. 希望休リスト
        headers = ["従業員CD", "日付", "休暇区分"]
        sht_name5 = "希望休リスト"
        data = [ ]

        for worker_code, date_dict in holidays.items():
            for date, hol in date_dict.items():
                if(type(date) == datetime.datetime and hol != ""):
                    data.append([worker_code, date.strftime("%Y/%m/%d"), hol])

        dfs[sht_name5] = pd.DataFrame(data, columns=headers)

        #SHEET6. 拘束時刻テーブル
        sht_name6 = "拘束時刻テーブル"
        data = [ ]
        
        data.append(["従業員CD", "従業員"] + [start + datetime.timedelta(i) for i in range(nday)] )
        data.append(["", ""] + ["開始-終了" for i in range(nday)])

        for worker_code, _ in skills.items():
            worker_name = worker_names[worker_code]
            if not worker_code in holidays:
                continue
            res = [worker_code, worker_name] + ["" for i in range(nday)]
            for day in range(nday):
                date  = start + datetime.timedelta(day)
                holi  = holidays[worker_code][day]
                other = otherworks[worker_code][day]

                if(worker_code in ignore and other != ""):                    
                    res[1 + 1 + day] = other

                for route_code, date2time in route.items():                
                    times = date2time[day]
                    start_time, end_time = times[0][0] + ":" + times[0][1], times[1][0] + ":" + times[1][1]                    
                    var = self.get_deltavar(worker_code, route_code, day)
                    if(var is not None):
                        if(self.getSol(var) > 0.1):
                            res[1 + 1 + day] = f"{start_time}-{end_time}"
            data.append(res)
        dfs[sht_name6] = pd.DataFrame(data)

        #Excel Wrter
        with pd.ExcelWriter(filepath) as writer:
            for sht, df in dfs.items():
                df.to_excel(writer, sheet_name=sht, index=False)            

        # ワークブックをロード
        wb = load_workbook(filepath)
        ws = wb["シフトテーブル"]

        # セルに色を付ける（例：2行目、1列目（A2セル）に黄色を適用）
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for irow, (worker_code, routes) in enumerate(skills.items()):
            if not worker_code in holidays:
                continue
            for day in range(nday): 
                icol = day + 1
                holi = holidays[worker_code][day]               
                for route_code, skill_level in routes.items():
                    var = self.get_deltavar(worker_code, route_code, day)
                    if(var is not None):
                        if(self.getSol(var) > 0.1 and holi):
                            ws.cell(irow + 2, icol + 1).fill = fill
        # 変更を保存
        wb.save(filepath)
        return 
   
    #解を保存
    def save(self):
        
        self.bestSol = dict()
        for key, var in self.vars.items():
            self.bestSol[key] = self.getSol(var)

        return 
    
    #日付などの情報を集計する
    def accept_master(self) -> None:
        #変数定義
        params : dict        = self.params
        nday   : int         = params[constant.param_nday]["value"]
        start  : datetime    = params[constant.param_targetmonth]["value"]
        #月のリストを作成
        self.month_list = list(set([ utils.to_month( start + datetime.timedelta(i-1) ) for i in range(nday) ] ) )
        #計画の開始日を取得
        self.offset_day = 0
        while(True):
            dt = start + datetime.timedelta(self.offset_day)
            if(dt.strftime('%A') == 'Monday'):
                break
            self.offset_day -= 1
        self.offset_day -= 13

        #計画の月初を取得
        self.offset_month = 0
        while(True):
            dt = start + datetime.timedelta(self.offset_month)
            if(dt.day == 1):
                break
            self.offset_month -= 1
        
        #当月の残り日数
        self.nday_this_month = 0
        while(True):
            dt = start + datetime.timedelta(self.nday_this_month)
            if(dt.month != start.month):
                break
            self.nday_this_month += 1
        
        #年度初めの日付を取得
        self.fiscal_year_start = utils.get_fiscal_year_start(start)

        # 日付とインデックスを対応づける
        for i in range(nday):
            date = start + datetime.timedelta(i)
            for key in self.master.holiday.keys():
                self.master.holiday[key][i] = self.master.holiday[key].get(date, "")
            for key in self.master.otherwork.keys():
                self.master.otherwork[key][i] = self.master.otherwork[key].get(date, "")
            for key in self.master.routeAssign.keys():
                self.master.routeAssign[key][i] = self.master.routeAssign[key].get(date, 0)
        for i in range(-1, nday):
            date = start + datetime.timedelta(i)
            for key in self.master.route.keys():
                self.master.route[key][i] = self.master.route[key].get(date, None)
        for i in range(-1, nday):
            date = start + datetime.timedelta(i)
            for key in self.master.runtime.keys():
                self.master.runtime[key][i] = self.master.runtime[key].get(date, 0)
        
        for i in range(self.offset_month, 0):
            date = start + datetime.timedelta(i)                
            for key in self.master.workedRouteRaw.keys():
                self.master.workedRouteRaw[key][i] = self.master.workedRouteRaw[key].get(date, "")

        # 他業務をスキルに追加する
        for worker_name, _ in self.master.skill.items():
            worker_other_dict = self.master.otherwork.get(worker_name, { })
            for _, other_working in worker_other_dict.items():
                if(other_working != ""):
                    self.master.skill[worker_name][other_working] = -1 #配車業務以外は負の値
        #日々実績
        for i in range(-30, 0):
            date = start + datetime.timedelta(i)
            for key in self.master.workedtimeRawDaily.keys():
                self.master.workedtimeRawDaily[key][i] = self.master.workedtimeRawDaily[key].get(date, [0 ,0])

        #法廷休日判定
        self.is_legal_holiday = dict()
        japanese_holidays = JapaneseHolidays()
        for i in range(-13, nday):
            date = start + datetime.timedelta(i)
            is_holiday = japanese_holidays.get_holiday_name(date)            
            if(is_holiday is not None or date.strftime('%A') == 'Sunday'):
                self.is_legal_holiday[i] = 1
        return 

    #最適化
    def optimize(self) -> cp_model.cp_model_pb2.CpSolverStatus.ValueType: 
        
        print("最適化開始")
        params = self.params
        
        status = cp_model.INFEASIBLE

        #求解できたか
        def resolved(status) -> bool:
            return status == cp_model.FEASIBLE or status == cp_model.OPTIMAL

        max_level = max([ params[constant.param_holiday_const_name]["value"], params[constant.param_pulic_holiday_const_name]["value"], 
                        params[constant.param_worktime_year_const_name]["value"], params[constant.param_worktime_month_const_name]["value"], 
                        params[constant.param_resttime_const_name]["value"], params[constant.param_dayave_const_name]["value"], 
                        params[constant.param_weekave_const_name]["value"] ])

        while not resolved(status) and self.relax_level <= max_level :
            #初期化
            self.init()
            #変数生成
            self.create_variables()
            #目的関数生成
            self.set_objective()
            #制約式生成
            self.create_constraints()

            #最適化実行
            if(self.solver_engin_is == "scip"):
                self.prob.optimize()
            else:
                #ソルバーインスタンス生成
                self.solver = cp_model.CpSolver()
                #各種パラメータを調整
                self.solver.parameters.log_search_progress = True
                self.solver.parameters.max_time_in_seconds = params[constant.param_timelimit]["value"]
                #self.solver.parameters.search_branching = cp_model.PORTFOLIO_SEARCH
                #self.solver.parameters.use_erwa_heuristic = True
                self.solver.parameters.symmetry_level = 3
                self.solver.parameters.linearization_level = 0  # 線形化のレベル設定
                self.solver.parameters.use_lns = True  # ラージ・ネイバー・サーチの使用

                # 詳細なヒューリスティックの設定例
                #self.solver.parameters.search_branching = cp_model.FIXED_SEARCH  # 固定探索
                #self.solver.parameters.randomize_search = True
                #self.solver.parameters.optimize_with_core = True  # コア最適化の使用
                #self.solver.parameters.linearization_level = 3  # 線形化レベルの設定
                
                
                self.solver.parameters.search_branching = cp_model.PORTFOLIO_SEARCH
                self.solver.parameters.use_erwa_heuristic = True
                #求解
                status = self.solver.Solve(self.prob)

            if(not resolved(status)):
                self.logger.info(f"最適化失敗 : 緩和レベル{self.relax_level}")

            self.relax_level += 1
        if(status == cp_model.FEASIBLE or status == cp_model.OPTIMAL) :    
            self.save()
        return status

    #目的関数生成
    def set_objective(self) -> None:
        obj = 0
        for key, var in self.vars.items():
            obj += self.coeffs[key] * var
        if(self.solver_engin_is == "scip"):
            self.prob.setObjective(obj, "maximize")
        else:
            self.prob.Maximize(obj)

    ##########################
    ######### 変数 ###########
    ##########################

    #モデルに変数を追加(vtype = C(ontinuos) or B(inary) or I(nteger) -> scipでのみ参照)
    def addVariable(self, key : str, lb : float, ub : float, vtype : str = "I"):
        if(self.solver_engin_is == "scip"):
            var = self.prob.addVar(key, vtype ,lb, ub)
        else:
            var = self.prob.NewIntVar(lb, ub, key)
        return var
    
    #変数を生成
    def create_variables(self) -> None:

        print("create variables...", end = "")
        #基礎変数(ワーカがどのルートどの日付割付くか）の生成
        self.create_deltavars()
        #残業変数の生成
        self.create_overtimevars()
        #同一ルート判定変数の生成
        self.create_contvars()
        #勤務日数変数の生成
        self.create_workingdayvars()
        #休日買取判定変数の生成
        self.create_returnholidayvars()
        #ワーカのルート割り付け判定変数の生成
        self.create_routevars()
        #時間ルールが守られているか判定する変数の生成
        self.create_timerulesvars()
        #ルートモードに関する変数の生成
        self.create_routemodevars()
        #稼働週のランクに関する変数の生成
        self.create_workrank()

        print("complete")

        return 
    #基礎変数の生成
    def create_deltavars(self) -> None:
        master    : masterClass = self.master
        nday      : int         = self.params[constant.param_nday]["value"]
        offset    : int         = self.offset_day
        skill     : dict        = master.skill
        holiday   : dict        = master.holiday
        otherwork : dict        = master.otherwork
        ignore    : set          = master.ignores

        for worker_name, routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for route_code, skill_level in routes.items():
                for day in range(offset, nday):
                    key = f"{worker_name},{route_code},{day},delta"
                    if(skill_level == 0):
                        continue
                    # 負であれば配車業務外
                    if(skill_level < 0 and route_code == otherwork[worker_name].get(day, "")):
                        lb, ub = 1, 1
                    else:
                        lb, ub = 0, 1
                    self.vars[key] = self.addVariable(key, lb, ub, "B")
                    if(0 <= day):
                        self.coeffs[key] = skill_level - 10000.0 if holiday[worker_name][day] and self.relax_level < 3 else 0.0
                    else:
                        self.coeffs[key] = 0

        return
    #残業変数の生成
    def create_overtimevars(self) -> None:
        master     : masterClass = self.master
        nday       : int         = self.params[constant.param_nday]["value"]
        month_list : list        = self.month_list
        skill      : dict        = master.skill
        holiday    : dict        = master.holiday
        ignore     : set         = master.ignores

        for worker_name, routes in skill.items():    
            if not worker_name in holiday or worker_name in ignore:
                continue
            for month in month_list:
                key = f"{worker_name},{month},overtime"
                self.vars[key] = self.addVariable(key, 0, 100, "C")
                self.coeffs[key] = -10.0
    #同一ルート判定変数の生成
    def create_contvars(self) -> None:
        master  : masterClass = self.master
        nday    : int         = self.params[constant.param_nday]["value"]
        skill   : dict        = master.skill
        holiday : dict        = master.holiday
        ignore  : set         = master.ignores
        
        for worker_name, routes in skill.items():    
            if not worker_name in holiday or worker_name in ignore:
                continue
            for day in range(nday):
                key = f"{worker_name},{day},cont"
                self.vars[key] = self.addVariable(key, 0, 4, "C")
                self.coeffs[key] = -50.0
        return
    #勤務日数変数の生成
    def create_routevars(self) -> None:
        master  : masterClass = self.master
        skill   : dict        = master.skill
        holiday : dict        = master.holiday
        ignore  : set         = master.ignores

        for worker_name , routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for route_code, skill_level in routes.items():
                key = f"{worker_name},{route_code},route"
                self.vars[key] = self.addVariable(key, 0, 1, "B")
                self.coeffs[key] = 0.0
        return         
    #ワーカのルート割り付け判定変数の生成
    def create_workingdayvars(self) -> None:
        master : masterClass = self.master
        nday   : int         = self.params[constant.param_nday]["value"]

        key = "workingday,min"
        self.vars[key] = self.addVariable(key, 0, nday + 7, "I")
        self.coeffs[key] = 1000
        key = "workingday,max"
        self.vars[key] = self.addVariable(key, 0, nday + 7, "I")
        self.coeffs[key] = -1000
    #休日買取判定変数の生成
    def create_returnholidayvars(self) -> None:
        master     : masterClass = self.master
        nday       : int         = self.params[constant.param_nday]["value"]
        month_list : list        = self.month_list
        skill      : dict        = master.skill
        holiday    : dict        = master.holiday
        ignore     : set         = master.ignores

        for worker_name, routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for month in month_list:
                key = f"{worker_name},{month},returnholiday"
                self.vars[key] = self.addVariable(key, 0, 1, "B")
                self.coeffs[key] = -100
        return 
    #時間ルールが守られているか判定する変数の生成
    def create_timerulesvars(self) -> None:
        master    : masterClass = self.master
        nday      : int         = self.params[constant.param_nday]["value"]
        offset    : int         = self.offset_day
        skill     : dict        = master.skill
        holiday   : dict        = master.holiday
        otherwork : dict        = master.otherwork
        ignore    : set          = master.ignores

        for worker_name, routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for day in range(-1, nday):
                key = f"{worker_name},{day},timerule"                          
                self.vars[key] = self.addVariable(key, 0, 1, "B")
                self.coeffs[key] = 0
                
        return
    #ルートモード変数の生成
    def create_routemodevars(self) -> None:
        master    : masterClass = self.master
        nday      : int         = self.params[constant.param_nday]["value"]
        offset    : int         = self.offset_day
        skill     : dict        = master.skill
        holiday   : dict        = master.holiday
        otherwork : dict        = master.otherwork
        ignore    : set          = master.ignores

        for worker_name, routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for route_code, skill_level in routes.items():
                for day in range(offset, nday):
                    key = f"{worker_name},{route_code},{day},mode"
                    if(skill_level == 0):
                        continue
                    lb, ub = 0, 1
                    self.vars[key] = self.addVariable(key, lb, ub, "B")
                    self.coeffs[key] = 0

        return
    #稼働週のランクに関する変数の生成
    def create_workrank(self) -> None:
        master    : masterClass = self.master
        nday      : int         = self.params[constant.param_nday]["value"]
        offset    : int         = self.offset_day
        skill     : dict        = master.skill
        holiday   : dict        = master.holiday
        otherwork : dict        = master.otherwork
        ignore    : set          = master.ignores

        for worker_name, routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for day in range(offset, nday):
                lb, ub = 0, nday
                key = f'{worker_name},{day},rank'
                self.vars[key] = self.addVariable(key, lb, ub, "I")
                self.coeffs[key] = 0

                lb, ub = 0, 1
                key = f'{worker_name},{day},rankup'
                self.vars[key] = self.addVariable(key, lb, ub, "B")
                self.coeffs[key] = 0
        return 

    #変数取得
    def getVar(self, key : str):
        if(key in self.vars):
            return self.vars[key]
        else:
            return None
    #基礎変数(ワーカがどのルートどの日付割付くか）の生成
    def get_deltavar(self, worker_name, route_code, day):
        key = f"{worker_name},{route_code},{day},delta"
        return self.getVar(key)
    #残業変数の生成
    def get_overtime(self, worker_name, month):
        key = f"{worker_name},{month},overtime"
        return self.getVar(key)
    #同一ルート判定変数の生成
    def get_contvar(self, worker_name, day):
        key = f"{worker_name},{day},cont"
        return self.getVar(key)
    #勤務日数変数の生成
    def get_workingdayvar(self, minmax):
        key = f"workingday,{minmax}"
        return self.getVar(key)
    #休日買取判定変数の生成
    def get_returnholidayvar(self, worker_name, month):
        key = f"{worker_name},{month},returnholiday"
        return self.getVar(key)
    #ワーカのルート割り付け判定変数の生成
    def get_routevar(self, worker_name, route):
        key = f"{worker_name},{route},route"
        return self.getVar(key)
    #時間ルールの判定変数の取得
    def get_timerulevar(self, worker_name, day):
        key = f"{worker_name},{day},timerule"
        return self.getVar(key)
    #ルートモードに関する変数の取得
    def get_routemodevar(self, worker_name, route_code, day):
        key = f"{worker_name},{route_code},{day},mode"
        return self.getVar(key)
    def get_rankvar(self, worker_name, day):
        key = f"{worker_name},{day},rank"
        return self.getVar(key)
    def get_rankupvar(self, worker_name, day):
        key = f"{worker_name},{day},rankup"
        return self.getVar(key)
    ##########################
    ######## 制約式 ###########
    ##########################

    #制約違反しているか
    def checkConstraintViolation(self, lhs, rhs):
        val_l, val_r = 0, 0
        for var in lhs:
            val_l += self.bestSol.get(str(var), 0)
        for var in rhs:
            val_r += var
        if(val_l > val_r):
            return True
        else:
            return False

    #モデルへ制約式を追加
    def addConstraint(self, lhs : list, rhs : list, type : str = "", is_relax : bool = False, name : str = "", keys : list = None) -> None:
        if(len(lhs) == 0 or len(rhs) == 0):
            return
        
        self.constraints.append({"is_relax" : is_relax, "lhs" : lhs, "rhs" : rhs, "type" : "e", "name" : name, "keys" : keys})

        if(is_relax):
            return 

        if(self.solver_engin_is == "scip"):
            if(type != "e"):
                self.prob.addCons(sum(lhs) <= sum(rhs))
            else:
                self.prob.addCons(sum(lhs) == sum(rhs))
        else:
            if(type != "e"):
                self.prob.Add(sum(lhs) <= sum(rhs))
            else:
                self.prob.Add(sum(lhs) == sum(rhs))        
        return
    
    #制約式を生成
    def create_constraints(self) -> None:
        print("create constraints...", end = "")

        ###ベースとなる制約式###
        
        #作業員のルートの割り付けに関する制約式を生成
        self.create_assign_constraint()

        #作業員が同一ルートに割りついているか判定する制約式を生成
        self.create_cont_constraint()

        #作業員の残業時間に関する制約式を生成
        self.create_overtime_constraint()

        #作業員の勤務日数を取得する制約式を生成
        self.create_workingday_constraints()

        #休日買取に関する制約式を生成
        self.create_returnholiday_constraints()

        #ルートモードに関する制約式を生成
        self.create_routemode_constraints()

        ###緩和の対象となる制約式###

        #作業員は週休2日とする制約式を生成 →休日労働は２週間に一度にする
        self.create_holiday_rule_constraint()        

        #業務終了時間と業務開始時間に間隔を設ける制約式を生成(あるルートの後にあるルートは割り付け不可)
        self.create_route2route_constraint()        

        #公休日取得に関する制約式を生成
        self.create_publicholiday_constraints()

        #1年間の拘束時間に関する制約式を生成
        self.create_worktime_year_constraints()

        #1ヶ月の拘束時間に関する制約式を生成
        self.create_worktime_month_constraints()

        #2日の1日平均に関する制約式を生成
        self.create_dayave_constraints()

        #2週の1週平均に関する制約式を生成
        self.create_weekave_constraints()

        #ランクごとの稼働開始時間に関する制約式を生成
        self.create_workrank_constraints()

        print("complete")
        return 
    
    #作業員のルートの割り付けに関する制約式を生成
    def create_assign_constraint(self) -> None:
        
        #変数定義
        master     : masterClass = self.master
        start_date : datetime    = self.params[constant.param_targetmonth]["value"]
        nday       : int         = self.params[constant.param_nday]["value"]
        offset_day : int         = self.offset_day
        skill      : dict        = master.skill
        route      : dict        = master.route        
        route_assignment : dict  = master.routeAssign
        otherwork  : set         = master.otherwork_name

        #実績部分の割り付け（ただし、制約違反の可能性もあるので一部の制約式は緩和）
        workedRouteRaw : dict =  master.workedRouteRaw
        for worker_name, routes in skill.items():
            if(not worker_name in workedRouteRaw):
                continue
            for day in range(offset_day, 0):
                route_code_raw = workedRouteRaw[worker_name].get(day, "")
                for route_code, _ in routes.items():
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        if(route_code_raw == route_code):
                            self.addConstraint([var], [1], "e")
                        else:
                            self.addConstraint([var], [0], "e")
        
        #計画部分の割り付け
        #各作業員は１日に１つのルートにしか割り当たらない
        for worker_name, routes in skill.items():
            for day in range(nday):
                linExpr = []
                for route_code, _ in routes.items():
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        linExpr.append(var)
                self.addConstraint(linExpr, [1])

        #各ルートには必ず１人の作業員が割り当たる
        start_dt = self.params[constant.param_targetmonth]["value"]      
        for route_code, _ in route.items():

            if(route_code in otherwork):
                continue

            for day in range(nday):
                linExpr = [0]

                assignment = route_assignment.get(route_code, {day : 1}).get(day, 1)                
                for worker_name, __ in skill.items():
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        linExpr.append(var)
                if(len(linExpr) > 1):
                    self.addConstraint(linExpr, [assignment], "e")    
        return 
    
    #作業員は週休2日とする制約式を生成 => 休日労働は２週間に1回に変更
    def create_holiday_rule_constraint(self) -> None:
        
        is_relax = 0 < self.params[constant.param_holiday_const_name]["value"] <= self.relax_level

        #変数を定義
        master     : masterClass   = self.master        
        nday       : int           = self.params[constant.param_nday]["value"]
        skill      : dict          = master.skill

        #休日労働は２週間に１度まで
        for worker_name, routes in skill.items():
            for day in range(-13, nday):
                linExpr = [] 
                for day2 in range(day, day + 14):
                    for route_code, skill_level in routes.items():
                        var = self.get_deltavar(worker_name, route_code, day2)
                        if(var is not None and self.is_legal_holiday.get(day2, 0)):
                            linExpr.append(var)

                self.addConstraint(linExpr, [1], type = "", is_relax = is_relax, name = "休日労働は２週間に１度までとする制約", keys = {"worker_code" : worker_name, "day" : day})
        return 
  
    #作業員が同一ルートに割りついているか判定する制約式を生成
    def create_cont_constraint(self) -> None:
        master : masterClass = self.master
        nday   : int         = self.params[constant.param_nday]["value"]
        skill  : dict        = master.skill

        for worker_name, routes in skill.items():
            rLinExpr = [ ]
            contvar = self.get_contvar(worker_name, 0)
            if(contvar is None):
                continue
            for route_code, skill_level in routes.items():
                if(skill_level < 0):
                    continue
                linExpr = [ ]
                route_var = self.get_routevar(worker_name, route_code)
                for day in range(0, nday):
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        linExpr.append(var)
                var = self.get_routevar(worker_name, route_code)
                if(var is not None):
                    if(len(linExpr) > 0):
                        if(self.solver_engin_is == "sat"):
                            self.prob.Add(sum(linExpr) != 0).OnlyEnforceIf(var)
                            self.prob.Add(sum(linExpr) == 0).OnlyEnforceIf(var.Not())
                        else:
                            self.addConstraint([var], linExpr, "")
                            self.addConstraint(linExpr, [len(linExpr)*var], "")
                    rLinExpr.append(var)
            rLinExpr.append(-contvar)
            self.addConstraint(rLinExpr, [1], "e") 
        return 

    #業務終了時間と業務開始時間に間隔を設ける制約式を生成(あるルートの後にあるルートは割り付け不可)    
    def create_route2route_constraint(self):

        is_relax = 0 < self.params[constant.param_resttime_const_name]["value"] <= self.relax_level

        master : masterClass = self.master
        nday   : int         = self.params[constant.param_nday]["value"]
        skill  : dict        = master.skill
        route  : dict        = master.route
        forbiddenes : list   = []    

        for worker_name, _ in skill.items():
            for day in range(-1, nday - 1):
                for route_code1, date2time1 in route.items():
                    for route_code2, date2time2 in route.items():
                        times, next_times = date2time1[day], date2time2[day+1]
                        
                        end_time = int(times[1][0]) * 60 + int(times[1][1])
                        next_start_time = int(next_times[0][0]) * 60 + int(next_times[0][1]) + 24 * 60
                        if(next_start_time - end_time < 9 * 60):
                            deltavar = self.get_deltavar(worker_name, route_code1, day )
                            next_deltavar = self.get_deltavar(worker_name, route_code2, day + 1)
                            if(deltavar is not None and next_deltavar is not None):                            
                                self.addConstraint([deltavar, next_deltavar], [1], type = "", is_relax = is_relax, name = "業務終了時間と業務開始時間に間隔を設ける", keys = {"worker_code" : worker_name, "day" : day})    

                            modevar = self.get_deltavar(worker_name, route_code1, day)
                            next_modevar = self.get_deltavar(worker_name, route_code2, day)
                            if(modevar is not None and next_modevar is not None):
                                self.addConstraint([modevar, next_modevar], [1], type = "", is_relax = is_relax, name = "業務終了時間と業務開始時間に間隔を設ける", keys = {"worker_code" : worker_name, "day" : day})

        return 
    
    #作業員の残業時間に関する制約式を生成
    def create_overtime_constraint(self):
        #変数定義
        master : masterClass       = self.master
        offset : int               = self.offset_day
        start  : datetime.datetime = self.params[constant.param_targetmonth]["value"]
        nday   : int               = self.params[constant.param_nday]["value"]        
        skill  : dict              = master.skill
        route  : dict              = master.route
        overworktime    : dict     = master.overworktime
        overworktimeRaw : dict     = master.overworktimeRaw
        month_list : list    = self.month_list
        date_obj = self.params[constant.param_targetmonth]["value"]

        # 月初の日付を計算
        for worker_name, skill in skill.items():
            for month in month_list:
                date = datetime.datetime.strptime(month, "%Y_%m_%d")
                overtimevar = self.get_overtime(worker_name, month)
                if(not worker_name in overworktime):
                    continue
                elif(overtimevar is None):
                    continue

                plan = overworktime[worker_name][date] if(worker_name in overworktime and date in overworktime[worker_name]) else 0
                raw  = overworktimeRaw[worker_name][date] if (worker_name in overworktimeRaw and date in overworktimeRaw[worker_name]) else 0
                target_time = (plan - raw) * 2

                linExpr = [-target_time]

                for route_code, date2times in route.items():                    
                    for day in range(nday):
                        times = date2times[day]
                        over_time0 = 8*60+30 - (int(times[0][0]) * 60 + int(times[0][1]))
                        over_time1 = (int(times[1][0]) * 60 + int(times[1][1])) - (17*60+30)

                        over_time0 *= (2/60)
                        over_time1 *= (2/60)
                        over_time0 = int(over_time0)
                        over_time1 = int(over_time1)

                        current_date = start + datetime.timedelta(day)
                        if(utils.to_month(current_date) != month):
                            continue
                        deltavar = self.get_deltavar(worker_name, route_code, day)
                        if(deltavar is not None):
                            if(0 < over_time0):
                                linExpr.append(over_time0 * deltavar)
                            if(0 < over_time1):
                                linExpr.append(over_time1 * deltavar)
                self.addConstraint(linExpr, [overtimevar])

        return
    
    #休日買取に関する制約式
    def create_returnholiday_constraints(self) -> None:

        master     : masterClass = self.master
        start      : datetime = self.params[constant.param_targetmonth]["value"]
        nday       : int = self.params[constant.param_nday]["value"]
        nholiday   : int = self.params[constant.param_publicholiday]["value"]
        offset     : int = self.offset_month #月初から判定するため（？）変更
        month_list : list = self.month_list
        skill      : dict = master.skill
        
        for worker_name, routes in skill.items():
            linExpr = []

            linExprDict = dict()
            for month in month_list :
                linExprDict[month] = list()

            for route_code, skill_level in routes.items():
                for day in range(offset, nday):
                    date = start + datetime.timedelta(day)
                    month = utils.to_month(date)
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        linExprDict[month].append(var)

            for month, linExpr in linExprDict.items():
                rvar = self.get_returnholidayvar(worker_name, month)
                if(rvar is not None):
                    self.prob.Add(sum(linExpr) <= nday - nholiday).OnlyEnforceIf(rvar.Not())
                    self.prob.Add(sum(linExpr) > nday - nholiday).OnlyEnforceIf(rvar)

        return 

    #ルートモードに関する制約式を生成
    def create_routemode_constraints(self) -> None:

        master    : masterClass = self.master
        nday      : int         = self.params[constant.param_nday]["value"]
        offset    : int         = self.offset_day
        skill     : dict        = master.skill
        holiday   : dict        = master.holiday
        otherwork : dict        = master.otherwork
        ignore    : set          = master.ignores

        id = 0

        for worker_name, routes in skill.items():
            if not worker_name in holiday or worker_name in ignore:
                continue
            for route_code, skill_level in routes.items():
                for day in range(offset, nday):
                    modevar = self.get_routemodevar(worker_name, route_code, day)
                    delvar = self.get_deltavar(worker_name, route_code, day)
                    if(delvar is not None and modevar is not None):
                        self.addConstraint([delvar], [modevar])

            for day in range(offset, nday - 1):
                linExpr = [ ]
                for route_code, skill_level in routes.items():
                    delvar = self.get_deltavar(worker_name, route_code, day + 1)
                    if(delvar is not None):
                        linExpr.append(delvar)
                for route_code, skill_level in routes.items():
                    modevar1 = self.get_routemodevar(worker_name, route_code, day)
                    modevar2 = self.get_routemodevar(worker_name, route_code, day + 1)

                    if(modevar1 is not None and modevar2 is not None):                    
                        self.addConstraint([modevar1, -modevar2], linExpr)

            for day in range(offset, nday):
                linExpr = [ ]
                for route_code, skill_level in routes.items():
                    modevar = self.get_routemodevar(worker_name, route_code, day)
                    if(modevar is not None):
                        linExpr.append(modevar)
                self.addConstraint(linExpr, [1], "e")

        return

    #作業員の勤務日数を取得する制約式を生成
    def create_workingday_constraints(self) -> None:

        master  : masterClass = self.master
        nday    : int         = self.params[constant.param_nday]["value"]
        skill   : dict        = master.skill
        holiday : dict        = master.holiday
        others  : dict        = master.otherwork

        for worker_name, routes in skill.items():
            if not worker_name in holiday:
                continue
            linExpr = []
            minvar, maxvar = self.get_workingdayvar("min"), self.get_workingdayvar("max")

            for route_code, skill_level in routes.items():
                for day in range(nday):
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        linExpr.append(var)

            all_zero = len([ 1 for route_code, skill_level in routes.items() if(skill_level > 0) ]) == 0
            if(not all_zero):
                self.addConstraint(linExpr, [maxvar])
                self.addConstraint([minvar], linExpr)

        return 
    
    #公休日取得に関する制約式を生成
    def create_publicholiday_constraints(self) -> None:

        is_relax = 0 < self.params[constant.param_pulic_holiday_const_name]["value"] <= self.relax_level

        #変数を定義
        master     : masterClass = self.master
        nday       : int         = self.params[constant.param_nday]["value"]
        nholiday   : int         = self.params[constant.param_publicholiday]["value"]
        offset     : int         = self.offset_month
        nday_month : int         = self.nday_this_month
        skill      : dict        = master.skill
        for worker_name, routes in skill.items():
            linExpr = []
            for route_code, skill_level in routes.items():
                for day in range(offset, nday_month):
                    var = self.get_deltavar(worker_name, route_code, day)
                    if(var is not None):
                        linExpr.append(var)
            if(not self.is_relax_log):
                self.addConstraint(linExpr, [nday_month - offset - nholiday], type = "", is_relax = is_relax, name = "公休日取得", keys = {"worker_code" : worker_name})            
        return

    #1年間の拘束時間に関する制約式を生成
    def create_worktime_year_constraints(self) -> None:

        is_relax = 0 < self.params[constant.param_worktime_year_const_name]["value"] <= self.relax_level

        #変数を定義
        master : masterClass = self.master
        start      : datetime = self.params[constant.param_targetmonth]["value"]
        nday : int = self.params[constant.param_nday]["value"]
        this_month = utils.to_month(start)
        skill      : dict = master.skill
        route  : dict        = master.route

        fiscal_year_start = utils.get_fiscal_year_start(start)
        
        # 関数を呼び出して結果を取得
        months = utils.get_months_in_period(start, nday)
        months2days = dict()        
        for m in months:
            months2days[m] = list()

        for day in range(0, nday):            
            month = utils.to_month(start + datetime.timedelta(day))
            month = datetime.datetime.strptime(month, "%Y_%m_%d")
            if(month in months):
                months2days[month].append(day)
        
        #１）1年の拘束時間：3,300時間以内        
        for worker_name, routes_skills in skill.items():            
            overworktimeRaw = master.overworktimeRaw.get(worker_name, {})
            linExpr = []
            nmonth = len(months)
            year_limit = 3300

            current_month = fiscal_year_start
            while(current_month <= months[-1]):
                year_limit -= overworktimeRaw.get(current_month, 0)
                current_month += relativedelta(months=1)
                nmonth += 1
            year_limit -= 242 * (12 - nmonth)
            
            for month, days in months2days.items():
                for route_code, _ in routes_skills.items():
                    if(not route_code in route):
                        continue
                    for day in days:                        
                        var = self.get_deltavar(worker_name, route_code, day)
                        times = route[route_code][day]
                        if(int(times[1][0]) < int(times[0][0])):
                            dt = 24 * 60 + int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                        else:
                            dt = int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))

                        if(var is not None):
                            linExpr.append(dt * var)

            self.addConstraint(linExpr, [year_limit * 60], type = "", is_relax = is_relax, name = "1年の拘束時間", keys = {"worker_code" : worker_name})
        return 

    #1ヶ月の拘束時間に関する制約式を生成
    def create_worktime_month_constraints(self) -> None:
        
        is_relax = 0 < self.params[constant.param_worktime_month_const_name]["value"] <= self.relax_level

        #変数を定義
        master : masterClass = self.master
        start      : datetime = self.params[constant.param_targetmonth]["value"]
        nday : int = self.params[constant.param_nday]["value"]
        this_month = utils.to_month(start)
        skill      : dict = master.skill
        route  : dict        = master.route

        # 関数を呼び出して結果を取得
        months = utils.get_months_in_period(start, nday)

        months2days = dict()        
        for m in months:
            months2days[m] = list()

        for day in range(0, nday):            
            month = utils.to_month(start + datetime.timedelta(day))
            month = datetime.datetime.strptime(month, "%Y_%m_%d")
            if(month in months):
                months2days[month].append(day)
        #２）1ケ月の拘束時間：310時間以内
        for worker_name, routes_skills in skill.items():
            overworktimeRaw = master.overworktimeRaw.get(worker_name, {})
            linExpr = []
            for month, days in months2days.items():
                month_limit = 310 - overworktimeRaw.get(month, 0)
                linExpr = []
                for route_code, _ in routes_skills.items():
                    if(not route_code in route):
                        continue
                    for day in days:
                        times = route[route_code][day]                    
                        if(int(times[1][0]) < int(times[0][0])):
                            dt = 24 * 60 + int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                        else:
                            dt = int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                        var = self.get_deltavar(worker_name, route_code, day)
                        if(var is not None):
                            linExpr.append(dt * var)
                self.addConstraint(linExpr, [month_limit * 60], type = "", is_relax = is_relax, name = "1ケ月の拘束時間", keys = {"worker_code" : worker_name, "day" : day})

        return 

    #2日の1日平均に関する制約式を生成
    def create_dayave_constraints(self) -> None:
        
        is_relax = 0 < self.params[constant.param_dayave_const_name]["value"] <= self.relax_level

        #変数を定義
        master     : masterClass = self.master
        start      : datetime    = self.params[constant.param_targetmonth]["value"]
        nday       : int         = self.params[constant.param_nday]["value"]
        this_month : str         = utils.to_month(start)
        skill      : dict        = master.skill
        route      : dict        = master.route
        raws       : dict        = master.workedtimeRawDaily
        runtimes   : dict        = master.runtime
 
        #５）5-1 1日週間の運転時間：2日平均1日の運転時間は、9時間以内、5-2 2週平均1週の運転時間は、44時間以内
        for worker_code, routes_skills in skill.items():
            for day in range(0, nday - 1):
                linExpr1, linExpr2 = [], []
                for route_code, skill_level in routes_skills.items():
                    if(not route_code in route):
                        continue

                    for i in range(-1, 1):
                        var = self.get_deltavar(worker_code, route_code, day + i)
                        if(raws[worker_code].get(day+i, None) is not None):
                            datetimes = raws[worker_code].get(day+i)
                            delta = math.ceil((datetimes[1] - datetimes[0]).total_seconds() / 60.0)
                            linExpr1.append(max(0, int(delta)))
                        elif(var is not None):
                            """
                            times = route[route_code][day + i]
                            if(int(times[1][0]) < int(times[0][0])):
                                dt = 24 * 60 + int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                            else:
                                dt = int(times[1][0]) * 60 + int(times[1][1]) - (int(times[0][0]) * 60 + int(times[0][1]))
                            """
                            dt = runtimes[route_code][day + i]
                            linExpr1.append(dt * var)
                    for i in range(0, 2):
                        var = self.get_deltavar(worker_code, route_code, day + i)
                        if(raws[worker_code].get(day+i, None) is not None):
                            datetimes = raws[worker_code].get(day+i)
                            delta = math.ceil((datetimes[1] - datetimes[0]).total_seconds() / 60.0)
                            linExpr2.append(max(0, int(delta)))
                        elif(var is not None):
                            dt = runtimes[route_code][day + i]
                            linExpr2.append(dt * var)

                rule_is_no_good1, rule_is_no_good2 = self.get_timerulevar(worker_code, day-1), self.get_timerulevar(worker_code, day)
                if(rule_is_no_good1 is not None and rule_is_no_good2 is not None):                    
                    self.prob.Add(sum(linExpr1) >= 9 * 60 * 2).OnlyEnforceIf(rule_is_no_good1)
                    self.prob.Add(sum(linExpr1) < 9 * 60 * 2).OnlyEnforceIf(rule_is_no_good1.Not())
                    self.prob.Add(sum(linExpr2) >= 9 * 60 * 2).OnlyEnforceIf(rule_is_no_good2)
                    self.prob.Add(sum(linExpr2) < 9 * 60 * 2).OnlyEnforceIf(rule_is_no_good2.Not())
                    self.addConstraint([rule_is_no_good1, rule_is_no_good2], [1], type = "", is_relax = is_relax, name = "2日平均1日の運転時間", keys = {"worker_code" : worker_code, "day" : day})

        return 

    #2週の1週平均に関する制約式を生成
    def create_weekave_constraints(self) -> None:
        
        is_relax = 0 < self.params[constant.param_weekave_const_name]["value"] <= self.relax_level

        #変数を定義
        master     : masterClass = self.master
        start      : datetime = self.params[constant.param_targetmonth]["value"]
        nday       : int = self.params[constant.param_nday]["value"]
        this_month : str = utils.to_month(start)
        skill      : dict = master.skill
        route      : dict        = master.route
        raws       : dict = master.workedtimeRawDaily
        runtimes   : dict = master.runtime

        for worker_code, routes_skills in skill.items():
            for day in range(-13, nday - 13):
                linExpr = []
                for route_code, skill_level in routes_skills.items():
                    if(not route_code in route):
                        continue

                    for i in range(0, 14):
                        if(day + i >= nday):
                            break

                        var = self.get_deltavar(worker_code, route_code, day + i)
                        if(raws[worker_code].get(day + i, None) is not None):
                            datetimes = raws[worker_code].get(day+i)
                            delta = math.ceil((datetimes[1] - datetimes[0]).total_seconds() / 60.0)
                            linExpr.append(max(0, int(delta)))                            
                        elif(var is not None):
                            dt = runtimes[route_code].get(day + i, 0) #日付がないときは０にする
                            linExpr.append(dt * var)
                self.addConstraint(linExpr, [44 * 2 * 7 * 60], type = "", is_relax = is_relax, name = "2週平均1週の運転時間", keys = {"worker_code" : worker_code, "day" : day})
        return 

    #ランクごとの稼働開始時間に関する制約式を生成
    def create_workrank_constraints(self):
        #変数を定義
        master     : masterClass = self.master
        start      : datetime = self.params[constant.param_targetmonth]["value"]
        nday       : int = self.params[constant.param_nday]["value"]
        difftime   : int = self.params[constant.param_diffstarttime]["value"]
        this_month : str = utils.to_month(start)
        skill      : dict = master.skill
        route      : dict        = master.route
        raws       : dict = master.workedtimeRawDaily
        runtimes   : dict = master.runtime
        ignores    : dict = master.ignores

        for worker_code, routes_skills in skill.items():
            for day in range(0, nday - 1):
                linExpr1, linExpr2 = [],[]
                for route_code, _ in route.items():
                    var1, var2 = self.get_deltavar(worker_code, route_code, day), self.get_deltavar(worker_code, route_code, day + 1)
                    if(var1 is not None and var2 is not None):
                        linExpr1.append(var1)
                        linExpr2.append(var2)

                rankvar1, rankvar2 = self.get_rankvar(worker_code, day), self.get_rankvar(worker_code, day + 1)
                rankupvar = self.get_rankupvar(worker_code, day)
                if(rankvar1 is not None and rankvar2 is not None):
                   self.prob.Add(sum(linExpr1) != sum(linExpr2)).OnlyEnforceIf(rankupvar)
                   self.prob.Add(sum(linExpr1) == sum(linExpr2)).OnlyEnforceIf(rankupvar.Not())

                   self.prob.Add(rankvar1 + 1 == rankvar2).OnlyEnforceIf(rankupvar)
                   self.prob.Add(rankvar1 == rankvar2).OnlyEnforceIf(rankupvar.Not())

        for worker_code, routes_skills in skill.items():
            
            if(worker_code in ignores):
                continue

            linExprs = []
            for day in range(nday):
                linExpr = []
                for route_code, date2time in route.items():
                    times = date2time[day]
                    start_time = int(times[0][0]) * 60 + int(times[0][1])
                    var = self.get_deltavar(worker_code, route_code, day)
                    if(var is not None):
                        linExpr.append(start_time * var)
                linExprs.append(linExpr)

            for day in range(nday - 1):
                for day2 in range(day+1, min(day+14, nday)):
                    linExpr1, linExpr2 = linExprs[day], linExprs[day2]
                    rankvar1, rankvar2 = self.get_rankvar(worker_code, day), self.get_rankvar(worker_code, day2)
                    
                    linExpr = linExpr1 + [-var for var in linExpr2]

                    self.addConstraint(linExpr, [difftime + 24 * 60 * (rankvar2 - rankvar1)])
                    self.addConstraint(linExpr, [difftime + 24 * 60 * (rankvar2 - rankvar1)])


        return
